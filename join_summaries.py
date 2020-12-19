"""
join_summaries.py
Author: Maggie Jacoby
Edited: 2020-12-18

This is used to summarize the completeness of each day/hub, and combine summaries for audio and images.
No inputs are needed when running the script.

Outputs summaries (csv) for all days/hubs, plus json with list of days above threshold (default 80%)

"""

# To Do: write occupancy part, print occcupancy to full df, copy "chosen days"
# write code to look at xlsx and create table of just days

import os
import sys
import csv
import argparse
import numpy as np
import pandas as pd
from glob import glob
from datetime import datetime, timedelta, time

from my_functions import *
import json

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, Rule


""" Run Parameters """
perc_threshold = 0.8
systems = ['H1-red', 'H2-red', 'H3-red', 'H4-red', 'H5-red', 'H6-black']
write_loc = make_storage_directory('/Users/maggie/Desktop/CompleteSummaries/2020-12-18')


def read_audio_summary(filepath):
    df = pd.read_csv(filepath, sep=' ', header=None)
    df.columns = ['hub', 'date', 'in', 'out', 'perc']
    df.index = df['date']
    df.drop(columns = df.columns.difference(['perc']), inplace=True)
    return df

def read_env_summary(filepath):
    df = pd.read_csv(filepath, sep=' ', header=None)
    df.columns = ['hub', 'date', 'in', 'out', 'perc']
    df.index = df['date']
    df.drop(columns = df.columns.difference(['perc']), inplace=True)
    return df

def read_img_summary(filepath, read_dark):
    df = pd.read_csv(filepath, sep=' ', header=0)
    read_cols = ['day', '%Capt', '%Dark'] if read_dark else ['day', '%Capt']
    col_names = ['date', 'perc', 'dark'] if read_dark else ['date', 'perc']
    df.drop(columns = df.columns.difference(read_cols), inplace=True)
    df.columns = col_names
    df.index = df['date']
    df.drop(columns = ['date'], inplace=True)
    return df


def format_xlsx(workbook, df, home_system, read_dark=False):
    ws = workbook.create_sheet(home_system, -1)

    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)
    ws['A1'] = ws['A2'].value
    ws.delete_rows(idx=2, amount=1)

    col_fin = get_column_letter(ws.max_column)
    col_second = get_column_letter(ws.max_column-1)
    n = int(len(ws['A']))

    full_range = f'B2:{col_second}{n}'
    grn1 = PatternFill(bgColor='00CCFFCC')

    full_cols_rule = CellIsRule(operator='greaterThanOrEqual', 
                formula=[perc_threshold], stopIfTrue=False, fill=grn1)

    ws.conditional_formatting.add(full_range, full_cols_rule)

    if not read_dark:
        date_range = f'A2:A{n}'
        final_range = f'{col_fin}2:{col_fin}{n}'
        blu1 = PatternFill(bgColor='0099CCFF')
        ylw1 = PatternFill(bgColor='00FFCC00')

        min_rule = CellIsRule(operator='greaterThanOrEqual',
                    formula=[perc_threshold], stopIfTrue=False, fill=blu1)

        date_col_rule = Rule(type="expression", dxf=DifferentialStyle(fill=ylw1))
        date_col_rule.formula = [f'${col_fin}2>={perc_threshold}']

        ws.conditional_formatting.add(final_range, min_rule)
        ws.conditional_formatting.add(date_range, date_col_rule)
    
    return workbook


def write_min_days(df):
    df_mods = df.drop(df.filter(regex='dark').columns, axis=1, inplace=False)
    df_mods['Minimum'] = df_mods.min(axis=1)

    days_above = list(df_mods.loc[df_mods['Minimum'] >= perc_threshold].index)
    days_above_threshold[home_system] = days_above

    full_write_loc = os.path.join(write_loc, f'all_days_above_{perc_threshold}.json')

    with open(full_write_loc, 'w') as file:
        file.write(json.dumps(days_above_threshold))
    print(f'File written to: {full_write_loc} for {len(days_above_threshold)} homes.')


def write_occ(filepath, dates):
    df = pd.read_csv(filepath, usecols=['Unnamed: 0', 'occupied'], sep=',', header=0)
    df.columns = ['timestamp', 'occupied']
    df.index = pd.to_datetime(df['timestamp'])
    df.drop(columns=['timestamp'], inplace=True)
    df['day'] = df.index.date
    occ_percents = {}
    for day in dates:
        day = datetime.strptime(day, '%Y-%m-%d').date()
        day_df = df.loc[df['day'] == day]
        if len(day_df) != 0:
            perc_total = day_df['occupied'].sum()/len(day_df)
        else:
            perc_total = 0.00
        occ_percents[day] = perc_total
    occ_df = pd.DataFrame.from_dict(occ_percents, orient='index', columns=['occupied'])
    occ_df = occ_df.round(decimals=2)
    return occ_df


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description="Description") 
    parser.add_argument('-read_dark','--read_dark', default='yes', type=str)
    parser.add_argument('-write_min', '--write_min', default='no', type=str)

    args = parser.parse_args()
    read_dark = True if args.read_dark == 'yes' else False
    write_min = True if args.write_min == 'yes' else False

    days_above_threshold = {}
    workbook = Workbook()

    for home_system in systems:

        print(home_system)
        H_num, color = home_system.split('-')
        
        path = os.path.join(f'/Users/maggie/Desktop/HPD_mobile_data/HPD-env/HPD_mobile-{H_num}/{home_system}')
        hubs = sorted(mylistdir(path, bit=f'{color.upper()[0]}S', end=False))

        dfs = []

        for hub in hubs:
            env_filepath = os.path.join(path, 'Summaries', f'{H_num}-{hub}-data-summary.txt')
            env_df = read_env_summary(env_filepath)
            env_df.columns = [f'{hub}_E']
            dfs.append(env_df)
            
            audio_filepath = os.path.join(path, 'Summaries', f'{H_num}-{hub}-audio-summary.txt')
            audio_df = read_audio_summary(audio_filepath)
            audio_df.columns = [f'{hub}_A']
            dfs.append(audio_df)

            img_filepath = os.path.join(path, 'Summaries', f'{H_num}-{hub}-img-summary.txt')
            img_df = read_img_summary(img_filepath, read_dark)
            img_df.columns = [f'{hub}_I', f'{hub}_I-dark'] if read_dark else [f'{hub}_I']
            dfs.append(img_df)

        df = pd.concat(dfs, axis=1)
        df = df.sort_index()
        df = df.fillna(0)

        if not read_dark:
            df['Minimum'] = df.min(axis=1)

        dates = list(df.index)
        occ_filepath = os.path.join(path, 'Summaries', f'{home_system}-Occupancy_df.csv')
        occ_df = write_occ(occ_filepath, dates)
        df['occ'] = occ_df.values

        workbook = format_xlsx(workbook=workbook, df=df, home_system=home_system, read_dark=read_dark)

        if write_min:
            write_min_days(df)



    workbook.save(os.path.join(write_loc, f'all_homes_highlight_wdark_test.xlsx'))
